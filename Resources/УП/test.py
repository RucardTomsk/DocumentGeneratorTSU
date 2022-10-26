import openpyxl

def get_dict_competencies(wb):
    dict_competencies = {}
    sheet = wb["Компетенции"]
    _row = 2
    key = ''
    while (sheet.cell(row = _row, column=5).value != "exit"):
        if sheet.cell(row = _row, column=5).value != None and sheet.cell(row = _row, column=5).value != '-':
            col_test = 1
            while(sheet.cell(row = _row, column=col_test).value == None):
                col_test+=1
            key = sheet.cell(row = _row, column=col_test).value
            dict_competencies[key] = [sheet.cell(row = _row, column=4).value,{}]
        else:
            if sheet.cell(row = _row, column=5).value == '-':
                col_test = 2
                while(sheet.cell(row = _row, column=col_test).value == None):
                    col_test+=1
                dict_competencies[key][1][sheet.cell(row = _row, column=col_test).value] = sheet.cell(row = _row, column=4).value
        _row+=1
    return dict_competencies

def get_rezult_str(str_mas_cod,dict_competencies):
	mas_str = str_mas_cod.split(";")
	for index in range(len(mas_str)):
		if mas_str[index][0] == ' ':
			mas_str[index] = mas_str[index][1:len(mas_str[index])]

	str1 = ''
	str2 = ''
	dict_competencies_final = {}
	for key in dict_competencies.keys():
		for key2 in dict_competencies[key][1].keys():
			dict_competencies_final[key2] = [dict_competencies[key][1][key2],key+' '+dict_competencies[key][0]]
	for key in mas_str:
		str2+= key+' '+ dict_competencies_final[key][0]+'\n'
		if not(dict_competencies_final[key][1] in str1):
			str1 += dict_competencies_final[key][1] + '\n'

	str1 = str1[:len(str1)-1]
	str2 = str2[:len(str2)-1]

	return str1,str2

def get_form_control(sheet,_row):
	final_str = ''
	mas_form_control = ['Экзамен','Зачет','Зачет с оценкой']
	mas = []
	for index in range(4,7):
		all_semester = sheet.cell(row = _row, column=index).value
		if all_semester != None:
			for semester in all_semester:
				mas.append("Семестр " + semester+ ', ' + mas_form_control[index-4])
	
	for str_f_m in sorted(mas,key = lambda s: int(s[8])):
		final_str+=str_f_m+'\n\t'

	return final_str[:len(final_str)-2]

def get_start_col(max_col,sheet)->int:
	for _col in range(1,max_col):
		if sheet.cell(row = 1, column=_col).value == "Курс 1":
			return _col

def get_flag_table_semestr(wb):
	sheet = wb["План"]
	if sheet.cell(row = 2, column=1).value == None:
		return True
	else:
		return False

def counter_semestr(wb):
    sheet = wb["Титул"]
    value = sheet.cell(row = 43, column=openpyxl.utils.column_index_from_string('C')).value
    term = int(value.split(':')[1][1])
    if get_flag_table_semestr(wb):
        return term*2
    else:
        return term

def get_sum_L(sheet,_row,_col,counter_semestr):
	sum_L = 0
	_col = _col +1
	for _ in range(counter_semestr):
		value = sheet.cell(row = _row, column=_col).value
		if value != None:
			sum_L+=float(value)
		_col+=10
	return str(sum_L)

def main():
    name_table = input("Введите полный путь к плану")
    wb = openpyxl.load_workbook(name_table)
    #str1, str2 =get_rezult_str("ИУК 3.1; ИУК 3.2; ИУК 3.3; ИОПК 6.2; ИПК 2.2",get_dict_competencies(wb))

    #print(str1)
    sheet = wb["План"]
    str1,str2 = get_rezult_str("ИУК 1.1; ИУК 3.1; ИУК 6.1; ИОПК 3.1",get_dict_competencies(wb))
    print(str1)
    print(str2)
    #print(get_form_control(sheet,13))
    #start_col = get_start_col(sheet.max_column,sheet)
    #counter_sem = counter_semestr(wb)
    #print(get_sum_L(sheet,8,start_col,counter_sem))

if __name__ == '__main__':
	main()

