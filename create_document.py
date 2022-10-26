import openpyxl
from docxtpl import DocxTemplate
import docx
import compose_doc

def read_word_fila(name_fila,code_name_d):
	doc = docx.Document(name_fila)
	all_paras = doc.paragraphs
	index_start = -1
	text = ""
	for text_index in range(len(all_paras)):
		if all_paras[text_index].text == code_name_d:
			index_start = text_index
			break

	if index_start != -1:
		for text_index_2 in range(index_start,len(all_paras)):
			if all_paras[text_index_2].text == "Тематический план:":
				index_start = text_index_2+1
				break

		for text_index_3 in range(index_start,len(all_paras)):
			if not("Тема" in all_paras[text_index_3].text):
				break
			text+='\t'+all_paras[text_index_3].text+'\n'

	if text == "":
		text = "Здесь должны быть темы"
	return text

def return_dict_kode(name_table):
	wb = openpyxl.load_workbook(name_table)
	sheet = wb["Компетенции"]
	rows = sheet.max_row
	dictionary1 = {}
	for row_column in range(2,rows):
		cell = sheet.cell(row = row_column, column=1)
		if cell.value != None:
			dictionary1.update({cell.value:sheet.cell(row = row_column, column=4).value})
		else:
			continue

	dictionary2 = {}
	for row_column in range(3,rows):
		cell = sheet.cell(row = row_column, column=2)
		if cell.value != None:
			dictionary2.update({cell.value:sheet.cell(row = row_column, column=4).value})
		else:
			continue

	return dictionary1,dictionary2

def return_string_kode(name_table,sheet,_row):
	str1 = ''
	str2 = ''
	mas = []
	dict1 , dict2 = return_dict_kode(name_table)
	all_komp = sheet.cell(row = _row, column=openpyxl.utils.column_index_from_string('AK')).value.replace(' ','').split(';')
	#all_komp[len(all_komp)-1] = all_komp[len(all_komp)-1] + ';'
	#for i in range(0,len(all_komp),2):
	#	komp = all_komp[i]+' '+all_komp[i+1][:len(all_komp[i+1])-1]
	#	s = komp[1:][:len(komp)-5]+'-'+komp[len(komp[:len(komp)-4]):][:len(komp)-5][1]
	#	mas.append('\t'+s + '. '+ dict1[s] + '\n')
	##	str2 +='\t'+komp+'. '+ dict2[komp] + '\n'
	#	#print(str2)
	#index = 1
	#while index < len(mas):
	#	if mas[index] in mas[ : index]:
	#		mas.pop(index)
	#	else:
	#		index += 1
	#for i in mas:
	#	str1+=i

	for i in all_komp:
		index = i.split(".")[0]
		if '11' in i or '10' in i:
			a = i[:-4]+' '+i[len(i)-4:]
			index = index[1:-2]+'-'+index[len(index)-2:]
		else:
			a = i[:-3]+' '+i[len(i)-3:]
			index = index[1:-1]+'-'+index[-1]
		if not(index in mas):
			mas.append(index)
		str2+='\t' + a +". " + dict2[a] + '\n'

	for i in mas:
		str1+= '\t' + i +". " + dict1[i] + '\n'
	return str1,str2

def return_string_MP(sheet,_row):
	str_mp = ''
	mas = []
	mas_text = ['Экзамен','Зачет','Зачет с оценкой']
	for i in range(4,7):
		all_semester = sheet.cell(row = _row, column=i).value
		if all_semester != None:
			for semester in all_semester:
				mas.append('\tСеместр '+semester+', '+ mas_text[i-4]+ '.\n')
		else:
			continue
	for i in sorted(mas,key = lambda s: s[8]):
		str_mp +=i

	return str_mp

def return_sum_L(sheet,_row):
	sum_L = 0
	col = 20
	for i in range(1,3):
		namber = sheet.cell(row = _row, column=col).value
		if namber != None:
			sum_L+= float(namber)
		col+=11
	return str(sum_L)

def return_sum_C(sheet,_row):
	sum_C = 0
	col = 23
	for i in range(1,3):
		namber = sheet.cell(row = _row, column=col).value
		if namber != None:
			sum_C+= float(namber)
		col+=11
	return str(sum_C)

def return_sum_P(sheet,_row):
	sum_P = 0
	col = 22
	for i in range(1,3):
		namber = sheet.cell(row = _row, column=col).value
		if namber != None:
			sum_P+= float(namber)
		col+=11
	return str(sum_P)

def return_sum_LB(sheet,_row):
	sum_LB = 0
	col = 21
	for i in range(1,3):
		namber = sheet.cell(row = _row, column=col).value
		if namber != None:
			sum_LB+= float(namber)
		col+=11
	return str(sum_LB)

def return_mas_doc(name_table,name_f):
	wb = openpyxl.load_workbook(name_table)
	sheet = wb["План"]
	rows = sheet.max_row
	cols = sheet.max_column
	global_mas = []
	for row_column in range(6,rows):
		flag = sheet.cell(row = row_column, column=openpyxl.utils.column_index_from_string('AJ')).value
		if flag != None:
			print(sheet.cell(row = row_column, column=3).value)
			mas = [sheet.cell(row = row_column, column=3).value,sheet.cell(row = row_column, column=2).value]
			str1,str2 = return_string_kode(name_table,sheet,row_column)
			#print(str1)
			#print(str2)
			mas.append(str1)
			mas.append(str2)
			#mas_test = ['Дисциплина относится к обязательной части образовательной программы.','Дисциплина относится к части образовательной программы, формируемой участниками образовательных отношений, является обязательной для изучения.','Дисциплина относится к части образовательной программы, формируемой участниками образовательных отношений, предлагается обучающимся на выбор.']
			mas_test = ['Дисциплина относится к обязательной части образовательной программы.','Дисциплина относится к части образовательной программы, формируемой участниками образовательных отношений, является обязательной для изучения.','Дисциплина относится к части образовательной программы, формируемой участниками образовательных отношений, предлагается обучающимся на выбор.']
			if 'О' in sheet.cell(row = row_column, column=2).value:
				mas.append(mas_test[0])
			elif ('ДВ' or 'ФТД') in sheet.cell(row = row_column, column=2).value:
				mas.append(mas_test[2])
			else:
				mas.append(mas_test[1])
			mas.append(return_string_MP(sheet,row_column))
			mas.append(sheet.cell(row = row_column, column=openpyxl.utils.column_index_from_string('H')).value)
			mas.append(sheet.cell(row = row_column, column=openpyxl.utils.column_index_from_string('K')).value)
			mas.append(return_sum_L(sheet,row_column))
			mas.append(return_sum_C(sheet,row_column))
			mas.append(return_sum_P(sheet,row_column))
			mas.append(return_sum_LB(sheet,row_column))
			name_predmet = mas[0]
			if len(name_predmet.split("*")) > 1:
				name_predmet = name_predmet.split("*")[0][:-1]
			mas.append(read_word_fila(name_f,mas[1]+' '+name_predmet))
			global_mas.append(mas)
		else:
			continue
	return global_mas

def main():
	name_table = input("Введите полный путь к плану")
	name_doc = input("Введите полный путь к шаблону")
	key = ['NAME','CODE','DEVELOP_GOAL','DEVELOP_REZULT','PLACE','LIGHTING_SEMES','ZE','K','L','C','P','LR',"CONTENT"]
	for document in return_mas_doc(name_table):
		doc = DocxTemplate(name_doc)
		contex = {}
		for i in range(len(document)):
			contex.update({key[i]:document[i]})
		doc.render(contex)
		if "*" in document[0]:
			doc.save("doc/"+document[0].replace('/','_').replace('*','_')+".docx")
		else:
			doc.save("doc/"+document[0].replace('/','_')+".docx")
		


	compose_doc.combine_all_docx("asd.docx",name_table.split('\\')[1].split(".plx")[0])
if __name__ == '__main__':
	main()






















